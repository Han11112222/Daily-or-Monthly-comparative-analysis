
import calendar
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# ─────────────────────────────────────────────
# 단위/환산 상수
# ─────────────────────────────────────────────
MJ_PER_NM3 = 42.563          # MJ / Nm3
MJ_TO_GJ = 1 / 1000.0        # 1 GJ = 1000 MJ


def mj_to_gj(x):
    try:
        return x * MJ_TO_GJ
    except Exception:
        return np.nan


def mj_to_m3(x):
    try:
        return x / MJ_PER_NM3
    except Exception:
        return np.nan


# ─────────────────────────────────────────────
# 기본 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="도시가스 공급량: 일/월 기온 기반 예측력 비교",
    layout="wide",
)


# ─────────────────────────────────────────────
# 표/출력 유틸
# ─────────────────────────────────────────────
def format_table_generic(df: pd.DataFrame, percent_cols=None):
    percent_cols = percent_cols or []
    out = df.copy()
    for c in out.columns:
        if c in percent_cols:
            out[c] = pd.to_numeric(out[c], errors="coerce")
            out[c] = out[c].map(lambda v: f"{v:.2%}" if pd.notna(v) else "")
        else:
            if out[c].dtype.kind in "if":
                out[c] = out[c].map(lambda v: f"{v:,.0f}" if pd.notna(v) else "")
    return out


def show_table_no_index(df: pd.DataFrame, height=420):
    st.dataframe(df, use_container_width=True, hide_index=True, height=height)


def _make_display_table_gj_m3(df: pd.DataFrame):
    """예상공급량(MJ)이 있으면 GJ/㎥ 컬럼을 추가해서 보여주기용 DF 반환"""
    out = df.copy()

    if "예상공급량(MJ)" in out.columns and "예상공급량(GJ)" not in out.columns:
        out["예상공급량(GJ)"] = out["예상공급량(MJ)"].apply(mj_to_gj)
    if "예상공급량(MJ)" in out.columns and "예상공급량(㎥)" not in out.columns:
        out["예상공급량(㎥)"] = out["예상공급량(MJ)"].apply(mj_to_m3)

    # 테이블에서는 MJ 컬럼은 숨기고 싶으면 여기서 drop해도 되지만,
    # "임의 삭제 금지" 조건 때문에 drop은 하지 않음.
    return out


# ─────────────────────────────────────────────
# 데이터 로딩
# ─────────────────────────────────────────────
@st.cache_data
def load_daily_data():
    """
    반환:
      df_model     : 공급량(MJ)와 평균기온 둘 다 있는 구간 (예측/R² 계산용)
      df_temp_all  : 평균기온만 있어도 되는 전체 구간 (매트릭스/히트맵용)
    """
    excel_path = Path(__file__).parent / "공급량(일일실적).xlsx"
    df_raw = pd.read_excel(excel_path)

    df_raw = df_raw[["일자", "공급량(MJ)", "공급량(M3)", "평균기온(℃)"]].copy()
    df_raw["일자"] = pd.to_datetime(df_raw["일자"])
    df_raw["연"] = df_raw["일자"].dt.year
    df_raw["월"] = df_raw["일자"].dt.month
    df_raw["일"] = df_raw["일자"].dt.day

    df_model = df_raw.dropna(subset=["공급량(MJ)", "평균기온(℃)"]).copy()
    df_temp_all = df_raw.dropna(subset=["평균기온(℃)"]).copy()

    return df_model, df_temp_all


@st.cache_data
def load_monthly_plan():
    excel_path = Path(__file__).parent / "월별계획.xlsx"
    df = pd.read_excel(excel_path)

    # 다양한 포맷을 흡수
    # 기대 포맷 예:
    #   연, 월, 계획(MJ) 혹은 월별계획(MJ) 등
    rename_map = {}
    for c in df.columns:
        if c in ["연도", "년도", "연"]:
            rename_map[c] = "연"
        if c in ["월"]:
            rename_map[c] = "월"
        if c in ["계획(MJ)", "월별계획(MJ)", "계획량(MJ)", "월계획(MJ)"]:
            rename_map[c] = "계획(MJ)"
        if c in ["계획(GJ)", "월별계획(GJ)", "계획량(GJ)", "월계획(GJ)"]:
            # 이미 GJ면 MJ로 변환해서 내부는 MJ로 통일
            rename_map[c] = "계획(GJ)"

    df = df.rename(columns=rename_map).copy()

    if "계획(MJ)" not in df.columns and "계획(GJ)" in df.columns:
        df["계획(MJ)"] = df["계획(GJ)"] * 1000.0

    # 연/월 보정
    if "연" not in df.columns:
        # fallback: 첫 컬럼이 연도일 수 있음
        df["연"] = pd.to_numeric(df.iloc[:, 0], errors="coerce")
    if "월" not in df.columns:
        df["월"] = pd.to_numeric(df.iloc[:, 1], errors="coerce")

    df["연"] = pd.to_numeric(df["연"], errors="coerce").astype("Int64")
    df["월"] = pd.to_numeric(df["월"], errors="coerce").astype("Int64")

    # 계획(MJ) 보정
    if "계획(MJ)" not in df.columns:
        # fallback: 마지막 컬럼이 계획일 수 있음
        last = df.columns[-1]
        df["계획(MJ)"] = pd.to_numeric(df[last], errors="coerce")

    df["계획(MJ)"] = pd.to_numeric(df["계획(MJ)"], errors="coerce")
    df = df.dropna(subset=["연", "월", "계획(MJ)"]).copy()

    return df


# ─────────────────────────────────────────────
# 일별 분배 로직
# ─────────────────────────────────────────────
def nth_weekday_of_month(d: pd.Timestamp):
    """
    d가 해당 월의 'n번째 요일'인지 계산.
    weekday_idx: 0=월 ... 6=일
    nth_dow: 1,2,3...
    """
    first = d.replace(day=1)
    weekday_idx = d.weekday()
    # 해당 월 1일부터 d일까지 중 같은 요일 몇 번째인지
    days = pd.date_range(first, d, freq="D")
    nth = (days.weekday == weekday_idx).sum()
    return weekday_idx, int(nth)


def make_daily_plan_table(df_daily: pd.DataFrame, target_year: int, target_month: int, recent_window: int = 3):
    """
    최근 N년(후보) 중 해당 월 실적이 있는 연도만 사용해서
    '평일1/평일2/주말' + 'n번째 요일' 기준 평균 패턴을 만들고,
    해당 월 계획(MJ)을 일별로 배분.
    """
    # 대상월 데이터(후보기간)
    years = sorted(df_daily["연"].dropna().unique().tolist())
    cand_years = [y for y in years if y < target_year]
    cand_years = cand_years[-recent_window:] if len(cand_years) >= recent_window else cand_years

    # 후보 중 대상월 실적 존재 연도만
    used_years = []
    for y in cand_years:
        sub = df_daily[(df_daily["연"] == y) & (df_daily["월"] == target_month)]
        if len(sub) > 0:
            used_years.append(y)

    # 대상월 날짜 생성
    last_day = calendar.monthrange(target_year, target_month)[1]
    dates = pd.date_range(f"{target_year}-{target_month:02d}-01", f"{target_year}-{target_month:02d}-{last_day}", freq="D")

    df_target = pd.DataFrame({"일자": dates})
    df_target["연"] = target_year
    df_target["월"] = target_month
    df_target["일"] = df_target["일자"].dt.day
    df_target["요일"] = df_target["일자"].dt.day_name()

    # 구분(평일1/평일2/주말)
    # 월=0 ... 일=6
    dow = df_target["일자"].dt.weekday
    df_target["구분"] = np.where(
        dow.isin([5, 6]),
        "주말/공휴일",
        np.where(dow.isin([0, 4]), "평일1(월·금)", "평일2(화·수·목)"),
    )

    # n번째 요일
    wd_info = df_target["일자"].apply(lambda d: nth_weekday_of_month(d))
    df_target["weekday_idx"] = wd_info.apply(lambda x: x[0])
    df_target["nth_dow"] = wd_info.apply(lambda x: x[1])

    # 최근N년 기반 raw 계산(구분 + nth_dow + weekday)
    raws = []
    for _, r in df_target.iterrows():
        # 같은 월, 같은 구분, 같은 weekday_idx, 같은 nth_dow 인 날들의 평균(최근N년)
        pool = []
        for y in used_years:
            subm = df_daily[(df_daily["연"] == y) & (df_daily["월"] == target_month)].copy()
            if len(subm) == 0:
                continue
            subm["weekday_idx"], subm["nth_dow"] = zip(*subm["일자"].apply(nth_weekday_of_month))
            subm["구분"] = np.where(
                subm["일자"].dt.weekday.isin([5, 6]),
                "주말/공휴일",
                np.where(subm["일자"].dt.weekday.isin([0, 4]), "평일1(월·금)", "평일2(화·수·목)"),
            )
            hit = subm[
                (subm["구분"] == r["구분"])
                & (subm["weekday_idx"] == r["weekday_idx"])
                & (subm["nth_dow"] == r["nth_dow"])
            ]
            if len(hit) > 0:
                pool.append(hit["공급량(MJ)"].mean())
        raw = float(np.nanmean(pool)) if len(pool) > 0 else np.nan
        raws.append(raw)

    df_target["최근N년_평균공급량(MJ)"] = raws

    # raw가 비어있으면 요일 평균으로 보정
    if df_target["최근N년_평균공급량(MJ)"].isna().any():
        # 요일평균(최근N년 대상월)
        tmp = []
        for y in used_years:
            subm = df_daily[(df_daily["연"] == y) & (df_daily["월"] == target_month)].copy()
            if len(subm) == 0:
                continue
            subm["weekday_idx"] = subm["일자"].dt.weekday
            tmp.append(subm[["weekday_idx", "공급량(MJ)"]])
        if len(tmp) > 0:
            tmp = pd.concat(tmp, ignore_index=True)
            weekday_mean = tmp.groupby("weekday_idx")["공급량(MJ)"].mean().to_dict()
        else:
            weekday_mean = {}

        df_target["최근N년_평균공급량(MJ)"] = df_target.apply(
            lambda r: weekday_mean.get(r["weekday_idx"], np.nan) if pd.isna(r["최근N년_평균공급량(MJ)"]) else r["최근N년_평균공급량(MJ)"],
            axis=1,
        )

    # 비율 계산
    raw_sum = df_target["최근N년_평균공급량(MJ)"].sum()
    df_target["일별비율"] = df_target["최근N년_평균공급량(MJ)"] / raw_sum if raw_sum else np.nan

    # 대상월 계획(MJ) 가져오기
    df_plan = load_monthly_plan()
    plan_row = df_plan[(df_plan["연"] == target_year) & (df_plan["월"] == target_month)]
    if len(plan_row) == 0:
        plan_total = np.nan
    else:
        plan_total = float(plan_row["계획(MJ)"].iloc[0])

    df_target["예상공급량(MJ)"] = df_target["일별비율"] * plan_total

    # 검증용: 최근N년 총공급량
    total_mj = 0.0
    for y in used_years:
        subm = df_daily[(df_daily["연"] == y) & (df_daily["월"] == target_month)]
        total_mj += float(subm["공급량(MJ)"].sum())
    df_target["최근N년_총공급량(MJ)"] = total_mj

    # 디버그 테이블
    df_debug = df_target[["일자", "요일", "일", "구분", "weekday_idx", "nth_dow", "최근N년_평균공급량(MJ)", "일별비율", "예상공급량(MJ)"]].copy()

    return df_target, used_years, cand_years, df_debug, plan_total


# ─────────────────────────────────────────────
# 엑셀 내보내기
# ─────────────────────────────────────────────
def _apply_excel_style(ws):
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")


def export_yearly_daily_plan_excel(df_daily_plan_year: pd.DataFrame, year: int):
    """
    연간 일별 계획 다운로드(Excel)
    - 기존 시트 구성 유지 + 마지막 시트에 누적계획량(예시 이미지 형태) 추가
    - GJ, ㎥ 모두 표시되도록 구성
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) 연간 계획 본문
        base = df_daily_plan_year.copy()
        base["예상공급량(GJ)"] = base["예상공급량(MJ)"].apply(mj_to_gj)
        base["예상공급량(㎥)"] = base["예상공급량(MJ)"].apply(mj_to_m3)

        base.to_excel(writer, index=False, sheet_name="연간")
        ws = writer.book["연간"]
        _apply_excel_style(ws)

        # 2) 누적현황(기준일 입력형) 시트 추가
        #    (예시 이미지처럼: 기준일, 일/월/연 목표/누적/진행률)
        ws2 = writer.book.create_sheet("누적현황")

        # 헤더/레이아웃
        ws2["A1"] = "기준일"
        ws2["B1"] = f"{year}-01-01"  # 기본값(사용자가 바꿔도 됨)

        headers = ["구분", "목표(GJ)", "누적(GJ)", "목표(㎥)", "누적(㎥)", "진행률(GJ)"]
        for j, h in enumerate(headers, start=1):
            ws2.cell(row=3, column=j, value=h)

        rows = ["일", "월", "연"]
        for i, r in enumerate(rows, start=4):
            ws2.cell(row=i, column=1, value=r)

        # 목표값: 기준일 기준 "해당일/해당월/해당연" 목표는
        # - 일: 해당일의 예상공급량
        # - 월: 해당월(1~말) 합계
        # - 연: 1~12월 합계
        #
        # 누적값:
        # - 일: 해당일 예상공급량
        # - 월: 해당월 1일~기준일 합계
        # - 연: 1/1~기준일 합계
        #
        # 진행률(GJ): 누적(GJ)/목표(GJ)

        # 기준일 셀
        기준일셀 = "B1"

        # 기준일의 연/월/일 추출 셀(숨김 계산용)
        ws2["H1"] = "연"
        ws2["I1"] = "월"
        ws2["J1"] = "일"
        ws2["H2"] = f"=YEAR({기준일셀})"
        ws2["I2"] = f"=MONTH({기준일셀})"
        ws2["J2"] = f"=DAY({기준일셀})"

        # 연간 시트 범위(연간 시트 컬럼 가정: A=일자, ... , 예상공급량(GJ)=?, 예상공급량(㎥)=?)
        # base.to_excel 결과 컬럼 순서 그대로를 썼으니, 컬럼 위치 찾아서 대응
        cols = list(base.columns)
        # "일자"가 A열
        col_date = cols.index("일자") + 1
        col_gj = cols.index("예상공급량(GJ)") + 1
        col_m3 = cols.index("예상공급량(㎥)") + 1

        def xl_col(n):
            s = ""
            while n:
                n, r = divmod(n - 1, 26)
                s = chr(65 + r) + s
            return s

        date_col_letter = xl_col(col_date)
        gj_col_letter = xl_col(col_gj)
        m3_col_letter = xl_col(col_m3)

        # 연간 시트 이름
        sh = "연간"

        # 일 목표/누적: 해당 날짜와 같은 행의 값
        # XLOOKUP 사용 가능(엑셀 365 기준). 호환성 위해 INDEX/MATCH로 작성.
        # 목표(GJ) (일)
        ws2["B4"] = (
            f"=IFERROR("
            f"INDEX({sh}!${gj_col_letter}:${gj_col_letter}, MATCH({기준일셀}, {sh}!${date_col_letter}:${date_col_letter}, 0)),"
            f'""'
            f")"
        )
        # 누적(GJ) (일) = 동일
        ws2["C4"] = ws2["B4"].value

        # 목표(㎥) (일)
        ws2["D4"] = (
            f"=IFERROR("
            f"INDEX({sh}!${m3_col_letter}:${m3_col_letter}, MATCH({기준일셀}, {sh}!${date_col_letter}:${date_col_letter}, 0)),"
            f'""'
            f")"
        )
        # 누적(㎥) (일)
        ws2["E4"] = ws2["D4"].value

        # 월 목표/누적: SUMIFS
        # 월 목표(GJ) = 해당월 전체 합
        ws2["B5"] = (
            f"=SUMIFS({sh}!${gj_col_letter}:${gj_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,$I$2,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<\"&EDATE(DATE($H$2,$I$2,1),1))"
        )
        # 월 누적(GJ) = 해당월 1일~기준일까지
        ws2["C5"] = (
            f"=SUMIFS({sh}!${gj_col_letter}:${gj_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,$I$2,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<=\"&{기준일셀})"
        )

        # 월 목표(㎥)
        ws2["D5"] = (
            f"=SUMIFS({sh}!${m3_col_letter}:${m3_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,$I$2,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<\"&EDATE(DATE($H$2,$I$2,1),1))"
        )
        # 월 누적(㎥)
        ws2["E5"] = (
            f"=SUMIFS({sh}!${m3_col_letter}:${m3_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,$I$2,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<=\"&{기준일셀})"
        )

        # 연 목표/누적
        ws2["B6"] = f"=SUM({sh}!${gj_col_letter}:${gj_col_letter})"
        ws2["C6"] = (
            f"=SUMIFS({sh}!${gj_col_letter}:${gj_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,1,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<=\"&{기준일셀})"
        )
        ws2["D6"] = f"=SUM({sh}!${m3_col_letter}:${m3_col_letter})"
        ws2["E6"] = (
            f"=SUMIFS({sh}!${m3_col_letter}:${m3_col_letter},"
            f"{sh}!${date_col_letter}:${date_col_letter},\">=\"&DATE($H$2,1,1),"
            f"{sh}!${date_col_letter}:${date_col_letter},\"<=\"&{기준일셀})"
        )

        # 진행률(GJ)
        ws2["F4"] = "=IFERROR(C4/B4,\"\")"
        ws2["F5"] = "=IFERROR(C5/B5,\"\")"
        ws2["F6"] = "=IFERROR(C6/B6,\"\")"

        # 스타일
        _apply_excel_style(ws2)

        # 표시 형식
        for r in range(4, 7):
            for c in [2, 3, 4, 5]:
                ws2.cell(row=r, column=c).number_format = "#,##0"
            ws2.cell(row=r, column=6).number_format = "0.00%"

        # 숨김(계산용)
        ws2.column_dimensions["H"].hidden = True
        ws2.column_dimensions["I"].hidden = True
        ws2.column_dimensions["J"].hidden = True

    output.seek(0)
    return output


# ─────────────────────────────────────────────
# TAB 1: Daily 공급량 분석
# ─────────────────────────────────────────────
def tab_daily_plan(df_daily: pd.DataFrame):
    st.title("도시가스 공급량 – 일별계획 예측")
    st.subheader("🗓️ Daily 공급량 분석 – 최근 N년 패턴 기반 일별 계획")

    st.markdown("### 📁 1. 월별계획 엑셀 업로드(XLSX) (없으면 폴더에서 자동 탐색)")
    upload = st.file_uploader("월별 계획 엑셀 업로드", type=["xlsx"])

    # 업로드가 없으면 repo의 월별계획.xlsx 사용
    # (load_monthly_plan에서 읽는 파일과 동일)
    if upload is not None:
        # 업로드 파일을 월별계획.xlsx로 임시 저장
        tmp_path = Path(__file__).parent / "_uploaded_monthly_plan.xlsx"
        tmp_path.write_bytes(upload.getbuffer())
        # load_monthly_plan이 읽는 파일명을 바꿀 수 없으니, 월별계획.xlsx로 덮어쓰기
        (Path(__file__).parent / "월별계획.xlsx").write_bytes(tmp_path.read_bytes())

    # 계획 연도/월 선택
    df_plan = load_monthly_plan()
    years = sorted(df_plan["연"].dropna().unique().tolist())
    months = list(range(1, 13))

    col1, col2 = st.columns(2)
    with col1:
        target_year = st.selectbox("계획 연도 선택", years, index=len(years) - 1 if len(years) else 0)
    with col2:
        target_month = st.selectbox("계획 월 선택", months, index=0)

    recent_window = st.slider("최근 몇 년 평균으로 비율을 계산할까?", 1, 6, 3)

    df_result, used_years, cand_years, df_debug, plan_total = make_daily_plan_table(
        df_daily=df_daily,
        target_year=int(target_year),
        target_month=int(target_month),
        recent_window=int(recent_window),
    )

    st.caption(
        f"최근 {recent_window}년 후보({cand_years[0] if cand_years else '-'}년 ~ {cand_years[-1] if cand_years else '-'}년) "
        f"{target_month}월 패턴으로 {target_year}년 {target_month}월 일별 계획을 계산. "
        f"(해당월 실적이 없는 연도는 자동 제외)"
    )
    st.markdown(f"- **실제 학습에 사용된 연도(해당월 실적 존재):** {used_years} (총 {len(used_years)}개)")

    if pd.isna(plan_total):
        st.error("대상 연/월의 월별 계획(MJ)을 찾지 못했어. 월별계획.xlsx의 연/월/계획 컬럼을 확인해줘.")
        return

    st.markdown(f"- **{target_year}년 {target_month}월 사업계획 제출 공급량 합계:** {mj_to_gj(plan_total):,.0f} GJ")

    st.markdown("### 🧩 일별 공급량 분배 기준")
    st.markdown(
        """
- 주말/공휴일/명절: **요일(토/일) + 그 달의 n번째** 기준 평균(공휴일/명절도 주말 패턴으로 묶음)  
- 평일: **평일1(월·금)** / **평일2(화·수·목)** 으로 구분  
  기본은 **요일 + 그 달의 n번째(1째 월요일, 2째 월요일...)** 기준 평균  
- 일부 케이스 데이터가 부족하면 **요일 평균**으로 보정  
- 마지막에 **일별비율 합계가 1이 되도록 정규화(raw / SUM(raw))**
        """
    )

    # ───────────────
    # 월별 계획(1~12월) & 연간 총량 표(상단 박스)
    # ───────────────
    st.markdown("### 📌 월별 계획량(1~12월) & 연간 총량")

    df_plan_year = df_plan[df_plan["연"] == int(target_year)].copy()
    pivot = df_plan_year.pivot_table(index="연", columns="월", values="계획(MJ)", aggfunc="sum")

    # 없을 수 있으니 방어
    if len(pivot) > 0:
        row = pivot.iloc[0].reindex(range(1, 13)).fillna(0.0)
        row_gj = row.apply(mj_to_gj)
        row_m3 = row.apply(mj_to_m3)

        year_total_mj = row.sum()
        year_total_gj = mj_to_gj(year_total_mj)
        year_total_m3 = mj_to_m3(year_total_mj)

        show = pd.DataFrame(
            {
                "구분": ["사업계획(월별 계획) - GJ", "사업계획(월별 계획) - ㎥"],
                **{f"{m}월": [row_gj.get(m, 0.0), row_m3.get(m, 0.0)] for m in range(1, 13)},
                "연간합계": [year_total_gj, year_total_m3],
            }
        )
        show2 = format_table_generic(show, percent_cols=[])
        show_table_no_index(show2, height=140)

    # ───────────────
    # 일별 계획 표
    # ───────────────
    st.markdown("### 📋 2. 일별 계획표(예상공급량: GJ/㎥ 동시 표시)")

    view = df_result.copy()
    view_with_total = view.copy()

    total_row = {
        "일자": "",
        "연": "",
        "월": "",
        "일": "",
        "요일": "",
        "weekday_idx": "",
        "nth_dow": "",
        "구분": "",
        "공휴일여부": False,
        "최근N년_평균공급량(MJ)": view["최근N년_평균공급량(MJ)"].sum(),
        "최근N년_총공급량(MJ)": view["최근N년_총공급량(MJ)"].sum(),
        "일별비율": view["일별비율"].sum(),
        "예상공급량(MJ)": view["예상공급량(MJ)"].sum(),
    }
    view_with_total = pd.concat([view, pd.DataFrame([total_row])], ignore_index=True)

    view_show = _make_display_table_gj_m3(view_with_total)
    view_show = format_table_generic(view_show, percent_cols=["일별비율"])
    show_table_no_index(view_show, height=520)

    with st.expander("🔎 (검증) 대상월 '1째 월요일/2째 월요일...' 계산 확인 (weekday_idx/nth_dow/raw/비율)"):
        dbg_disp = format_table_generic(df_debug.copy(), percent_cols=["일별비율"])
        show_table_no_index(dbg_disp, height=420)

    st.markdown("#### 📊 2. 일별 예상 공급량 & 비율 그래프(평일1/평일2/주말 분리)")

    w1_df = view[view["구분"] == "평일1(월·금)"].copy()
    w2_df = view[view["구분"] == "평일2(화·수·목)"].copy()
    wend_df = view[view["구분"] == "주말/공휴일"].copy()

    # --- Hover 값이 막대(일별)와 다르게 보이는 문제 방지 ---
    # (Plotly의 자동 SI 포맷/hover 표시가 헷갈릴 수 있어서, hovertemplate로 '일별 예상공급량(GJ/㎥)'을 고정 표기)
    _view_plot = view.sort_values("일").copy()
    _view_plot["_date_str"] = pd.to_datetime(_view_plot["일자"]).dt.strftime("%Y-%m-%d")
    _view_plot["_y_gj"] = _view_plot["예상공급량(MJ)"].apply(mj_to_gj)
    _view_plot["_y_m3"] = _view_plot["예상공급량(MJ)"].apply(mj_to_m3)

    def _make_bar_customdata(df_sub: pd.DataFrame):
        return np.stack(
            [
                df_sub["_y_gj"].to_numpy(),
                df_sub["_y_m3"].to_numpy(),
                df_sub["_date_str"].astype(str).to_numpy(),
                df_sub["요일"].astype(str).to_numpy(),
                df_sub["구분"].astype(str).to_numpy(),
            ],
            axis=-1,
        )

    fig = go.Figure()

    # 평일1/평일2/주말 분리 막대 (hover에 '일별 예상공급량'을 명시)
    _w1 = _view_plot[_view_plot["구분"] == "평일1(월·금)"].copy()
    _w2 = _view_plot[_view_plot["구분"] == "평일2(화·수·목)"].copy()
    _we = _view_plot[_view_plot["구분"] == "주말/공휴일"].copy()

    fig.add_bar(
        x=_w1["일"],
        y=_w1["_y_gj"],
        name="평일1(월·금) 예상공급량(GJ)",
        customdata=_make_bar_customdata(_w1),
        hovertemplate=(
            "<b>%{customdata[2]}</b> (%{customdata[3]})"
            "<br>구분: %{customdata[4]}"
            "<br>예상공급량: %{customdata[0]:,.0f} GJ"
            "<br>예상공급량: %{customdata[1]:,.0f} ㎥"
            "<extra></extra>"
        ),
    )
    fig.add_bar(
        x=_w2["일"],
        y=_w2["_y_gj"],
        name="평일2(화·수·목) 예상공급량(GJ)",
        customdata=_make_bar_customdata(_w2),
        hovertemplate=(
            "<b>%{customdata[2]}</b> (%{customdata[3]})"
            "<br>구분: %{customdata[4]}"
            "<br>예상공급량: %{customdata[0]:,.0f} GJ"
            "<br>예상공급량: %{customdata[1]:,.0f} ㎥"
            "<extra></extra>"
        ),
    )
    fig.add_bar(
        x=_we["일"],
        y=_we["_y_gj"],
        name="주말/공휴일 예상공급량(GJ)",
        customdata=_make_bar_customdata(_we),
        hovertemplate=(
            "<b>%{customdata[2]}</b> (%{customdata[3]})"
            "<br>구분: %{customdata[4]}"
            "<br>예상공급량: %{customdata[0]:,.0f} GJ"
            "<br>예상공급량: %{customdata[1]:,.0f} ㎥"
            "<extra></extra>"
        ),
    )

    # 일별비율 라인 (hover에 %로 고정)
    _line_cd = np.stack(
        [
            _view_plot["_date_str"].astype(str).to_numpy(),
            _view_plot["요일"].astype(str).to_numpy(),
        ],
        axis=-1,
    )
    fig.add_trace(
        go.Scatter(
            x=_view_plot["일"],
            y=_view_plot["일별비율"],
            mode="lines+markers",
            name=f"일별비율 (최근{len(used_years)}년 실제 사용)",
            yaxis="y2",
            customdata=_line_cd,
            hovertemplate=(
                "<b>%{customdata[0]}</b> (%{customdata[1]})"
                "<br>일별비율: %{y:.2%}"
                "<extra></extra>"
            ),
        )
    )

    fig.update_layout(
        title=(
            f"{target_year}년 {target_month}월 일별 공급량 계획 "
            f"(최근{recent_window}년 후보 중 실제 사용 {len(used_years)}년, {target_month}월 패턴 기반)"
        ),
        xaxis_title="일",
        yaxis=dict(title="예상 공급량 (GJ)", tickformat=","),
        yaxis2=dict(title="일별비율", overlaying="y", side="right", tickformat=".1%"),
        barmode="group",
        hovermode="x unified",
        margin=dict(l=20, r=20, t=60, b=40),
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("#### 🧾 4. 구분별 비중 요약(평일1/평일2/주말)")

    summary = (
        view.groupby("구분", as_index=False)
        .agg(일별비율합계=("일별비율", "sum"), 예상공급량_MJ=("예상공급량(MJ)", "sum"))
        .rename(columns={"예상공급량_MJ": "예상공급량(MJ)"})
    )
    summary["예상공급량(GJ)"] = summary["예상공급량(MJ)"].apply(mj_to_gj).round(0)
    summary["예상공급량(㎥)"] = summary["예상공급량(MJ)"].apply(mj_to_m3).round(0)

    total_row_sum = {
        "구분": "합계",
        "일별비율합계": summary["일별비율합계"].sum(),
        "예상공급량(MJ)": summary["예상공급량(MJ)"].sum(),
        "예상공급량(GJ)": summary["예상공급량(GJ)"].sum(),
        "예상공급량(㎥)": summary["예상공급량(㎥)"].sum(),
    }
    summary = pd.concat([summary, pd.DataFrame([total_row_sum])], ignore_index=True)

    summary_show = summary[["구분", "일별비율합계", "예상공급량(GJ)", "예상공급량(㎥)"]].copy()
    summary_show = format_table_generic(summary_show, percent_cols=["일별비율합계"])
    show_table_no_index(summary_show, height=220)

    # ───────────────
    # (기존) 최근N년 월 일별 실적 매트릭스
    # ───────────────
    st.markdown("#### 🧊 3. (검증) 최근 N년 대상월 일별 실적 공급량(GJ) 매트릭스")

    if len(used_years) >= 1:
        mats = []
        for y in used_years:
            sub = df_daily[(df_daily["연"] == y) & (df_daily["월"] == target_month)][["일", "공급량(MJ)"]].copy()
            sub = sub.sort_values("일")
            sub["공급량(GJ)"] = sub["공급량(MJ)"].apply(mj_to_gj)
            sub = sub.set_index("일")["공급량(GJ)"]
            mats.append(sub.rename(str(y)))
        mat = pd.concat(mats, axis=1).reindex(range(1, last_day + 1))

        fig_hm = go.Figure(
            data=go.Heatmap(
                z=mat.values,
                x=mat.columns.tolist(),
                y=mat.index.tolist(),
                colorbar=dict(title="GJ"),
                colorscale="RdBu_r",
            )
        )
        fig_hm.update_layout(
            title=f"최근 {len(used_years)}년 {target_month}월 일별 실적 공급량(GJ) 매트릭스",
            xaxis=dict(title="연도", type="category"),
            yaxis=dict(title="일", autorange="reversed"),
            margin=dict(l=40, r=40, t=60, b=40),
        )
        st.plotly_chart(fig_hm, use_container_width=False)

    # ───────────────
    # 5. 일일계획 다운로드(월간)
    # ───────────────
    st.markdown("### 📥 5. 일일계획 다운로드(월간)")
    # 월간 다운로드는 필요 시 추가 구현 가능(요청 범위 외)

    # ───────────────
    # 6. 일일계획 다운로드(연간)
    # ───────────────
    st.markdown("### 📁 6. 일일계획 다운로드(연간)")
    year_sel = st.selectbox("연간 계획 연도 선택", years, index=years.index(int(target_year)) if int(target_year) in years else 0)

    if st.button(f"📥 {year_sel}년 연간 일별공급계획 다운로드(Excel)"):
        # 선택 연도 12개월 각각 일별 계획 생성 후 concat
        df_all = []
        for m in range(1, 13):
            df_m, _, _, _, _ = make_daily_plan_table(df_daily=df_daily, target_year=int(year_sel), target_month=m, recent_window=int(recent_window))
            df_all.append(df_m)
        df_year = pd.concat(df_all, ignore_index=True)

        out = export_yearly_daily_plan_excel(df_year, int(year_sel))
        st.download_button(
            "⬇️ Excel 다운로드",
            data=out,
            file_name=f"{year_sel}년_연간_일별공급계획.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ─────────────────────────────────────────────
# TAB 2: Daily-Monthly 공급량 비교
# ─────────────────────────────────────────────
def tab_daily_monthly_compare(df_daily: pd.DataFrame, df_temp_all: pd.DataFrame):
    st.title("도시가스 공급량 – 일별 vs 월별 예측 검증")

    # 0. 상관도 분석(공급량 vs 주요 변수)
    st.markdown("## 📊 0. 상관도 분석 (공급량 vs 주요 변수)")

    # 상관분석용 데이터 준비(필요 컬럼만)
    df_corr = df_daily.copy()
    df_corr = df_corr.dropna(subset=["공급량(MJ)", "평균기온(℃)"]).copy()

    # 예시로 주요 변수(있는 것만)
    cols = [c for c in ["공급량(MJ)", "유효일수", "평균기온(℃)", "최저기온(℃)", "최고기온(℃)", "체감온도(℃)",
                        "총인구수(명)", "세대수(세대)", "인구순이동(명)", "고령인구수(명)", "소비자물가지수(%)", "청구전"] if c in df_corr.columns]

    corr = df_corr[cols].corr(numeric_only=True)

    # 정사각형(셀)로 보이도록: figure size/scale + aspect 고정
    fig_corr = go.Figure(
        data=go.Heatmap(
            z=corr.values,
            x=corr.columns,
            y=corr.index,
            colorbar=dict(title="상관계수"),
            colorscale="Blues",
            zmin=-1,
            zmax=1,
            text=np.round(corr.values, 2),
            texttemplate="%{text}",
        )
    )
    fig_corr.update_layout(
        width=1100,
        height=650,
        margin=dict(l=80, r=40, t=70, b=40),
    )
    fig_corr.update_yaxes(autorange="reversed")
    fig_corr.update_xaxes(tickangle=45)
    st.plotly_chart(fig_corr, use_container_width=False)

    # ───────────────
    # (요청) 가장 아래: G. 기온분석 – 일일 평균기온 히트맵
    # ───────────────
    st.markdown("## 🧊 G. 기온분석 – 일일 평균기온 히트맵")
    st.caption("기본은 공급량 데이터의 평균기온(℃)을 사용해. 필요하면 기온 파일만 별도로 업로드해서 볼 수도 있어.")

    temp_upload = st.file_uploader("일일기온 파일 업로드(XLSX) (선택)", type=["xlsx"], key="temp_upload_cmp")

    df_temp = df_temp_all.copy()
    if temp_upload is not None:
        try:
            tdf = pd.read_excel(temp_upload)
            # 기대: 일자, 평균기온(℃) 혹은 유사 컬럼
            # 최소 컬럼 정규화
            if "일자" not in tdf.columns:
                # 첫 컬럼이 날짜일 수 있음
                tdf = tdf.rename(columns={tdf.columns[0]: "일자"})
            # 평균기온 컬럼 찾기
            temp_col = None
            for c in tdf.columns:
                if "평균" in str(c) and ("기온" in str(c) or "온도" in str(c)):
                    temp_col = c
                    break
            if temp_col is None:
                # fallback
                temp_col = tdf.columns[1]
            tdf = tdf[["일자", temp_col]].copy()
            tdf = tdf.rename(columns={temp_col: "평균기온(℃)"})
            tdf["일자"] = pd.to_datetime(tdf["일자"])
            tdf["연"] = tdf["일자"].dt.year
            tdf["월"] = tdf["일자"].dt.month
            tdf["일"] = tdf["일자"].dt.day
            df_temp = tdf.dropna(subset=["평균기온(℃)"]).copy()
        except Exception:
            st.warning("기온 파일을 읽는 데 실패했어. 기본(공급량 데이터의 평균기온)으로 표시할게.")

    # 슬라이더: 연도 범위
    years_temp = sorted(df_temp["연"].dropna().unique().tolist())
    if len(years_temp) >= 2:
        y_min, y_max = years_temp[0], years_temp[-1]
        yr = st.slider("연도 범위", min_value=int(y_min), max_value=int(y_max), value=(int(y_min), int(y_max)))
    elif len(years_temp) == 1:
        yr = (int(years_temp[0]), int(years_temp[0]))
    else:
        st.info("기온 데이터가 없어.")
        return

    # 월 선택
    month_pick = st.selectbox("월 선택", list(range(1, 13)), index=0, format_func=lambda m: f"{m:02d} (January)" if m == 1 else f"{m:02d}")

    df_h = df_temp[(df_temp["연"] >= yr[0]) & (df_temp["연"] <= yr[1]) & (df_temp["월"] == int(month_pick))].copy()
    if len(df_h) == 0:
        st.info("선택 범위에 데이터가 없어.")
        return

    pivot = df_h.pivot_table(index="일", columns="연", values="평균기온(℃)", aggfunc="mean").reindex(range(1, 32))

    fig_t = go.Figure(
        data=go.Heatmap(
            z=pivot.values,
            x=pivot.columns.astype(str).tolist(),
            y=pivot.index.tolist(),
            colorbar=dict(title="℃"),
            colorscale="RdBu_r",
            zmid=0,
            text=np.round(pivot.values, 1),
            texttemplate="%{text}",
        )
    )
    fig_t.update_layout(
        title=f"{month_pick:02d}월 일일 평균기온 히트맵(선택연도 {len(pivot.columns)}개)",
        xaxis=dict(title="연도", type="category"),
        yaxis=dict(title="Day", autorange="reversed"),
        margin=dict(l=60, r=60, t=70, b=40),
        height=620,
    )
    st.plotly_chart(fig_t, use_container_width=True)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    df_daily, df_temp_all = load_daily_data()

    st.sidebar.markdown("### 좌측 탭 선택")
    tab = st.sidebar.radio(
        "",
        ["Daily 공급량 분석", "Daily·Monthly 공급량 비교"],
        index=0,
        format_func=lambda x: "🗓️ Daily 공급량 분석" if x == "Daily 공급량 분석" else "📊 Daily·Monthly 공급량 비교",
    )

    if tab == "Daily 공급량 분석":
        tab_daily_plan(df_daily=df_daily)
    else:
        tab_daily_monthly_compare(df_daily=df_daily, df_temp_all=df_temp_all)


if __name__ == "__main__":
    main()
